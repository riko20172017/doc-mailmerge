from openpyxl import load_workbook
from mailmerge import MailMerge, MailMergeOptions, OptionAutoUpdateFields, OptionKeepFields
import qrcode
from io import BytesIO
import tempfile
import os
from PIL import Image
from pathlib import Path
import warnings
from docx import Document
from docx.oxml.ns import qn
from lxml import etree
import copy
from datetime import datetime
import locale
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def convertMarktoWord(mark):
    match mark:
        case "5":
            return "отлично"
        case "4":
            return "хорошо"
        case "3":
            return "удовлетворительно"
        case "x":
            return "x"
        case _:
            return mark


def fix_encoding(text):
    """Исправляет кодировку Windows-1251"""
    if not text:
        return ""
    
    # Если строка содержит последовательности вида РўРµСЃР»РµРЅРєРѕ
    if re.search(r'Р[А-Я]', str(text)):
        try:
            # Пробуем декодировать из Windows-1251
            return str(text).encode('latin-1').decode('windows-1251')
        except:
            return str(text)
    return str(text)


def format_date_ru(date_str):
    # Парсим исходную дату

    # Словарь с русскими названиями месяцев в родительном падеже
    months = {
        1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
        5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
        9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
    }

    # Форматируем: день с ведущим нулём, месяц из словаря, год, слово "года"
    return f"{date_str.day:02d} {months[date_str.month]} {date_str.year} года"

mailmerge_options = MailMergeOptions(
    remove_empty_tables=False,
    auto_update_fields_on_open=OptionAutoUpdateFields.ALWAYS,
    # keep_fields=OptionKeepFields.NONE,
    # merge_if_fields=True,
    table_rows_replace_mode=False)


data_vidachi = "01 июля 2026 года"
gruppa = "ОЭВМ-5"
diploms = []
wb = load_workbook(r'C:\\Users\\ganz\\OneDrive\\Документы\\колледж onedrive\\ведомости\\' + gruppa + '.xlsx')
wb_common = load_workbook('utils.xlsx')
ws_common = wb_common["data"]
ws_data = wb["Сведения"]
ws_mark = wb["Оценки"]
common = {}

for row in ws_common.values:
    grupp_name = row[0]
    if grupp_name == gruppa:
        common = {
            "reshenie": row[2],
            "predsedatel": row[3],
            "vidacha": row[6],
            "svid-qual": row[16],
            "svid-prof": row[17],
        }
        # Вывод колонок A и C
for i, row in enumerate(ws_data.iter_rows(min_row=2, max_row=50, min_col=1, max_col=13, values_only=True)):
    fio = row[0]
    if (fio == None): # Если дойти до пустой строки, то остановить цикл
        break

    first, middle, last = fio.split(' ')
    dt_obj = row[2]
    birthdate = format_date_ru(row[2])
    pred_document = ""
    predsedatel = common["predsedatel"]
    svid_resh = format_date_ru(common["reshenie"])
    svid_qual = common["svid-qual"]
    svid_prof = common["svid-prof"]
    svid_reg_nom = row[11]
    svid_nom_obl = row[12]
    
    subjects = ""
    hours = ""
    marks = ""

    student_col = i+3
    for mark_col in ws_mark.iter_rows(min_row=2, max_row=100, min_col=0, max_col=student_col, values_only=True):
        subject = mark_col[0]
        hour = mark_col[1]
        mark = mark_col[student_col-1]

        if subject == None:
            break
        if mark == None:
            continue

        subject = str(subject)
        hour = str(hour)
        mark = str(mark)

        mark = convertMarktoWord(mark) # Преобразовать числовые оценки в "отлично", "хорошо" и тд.

        separator = "*"
   
        subjects += subject + separator

        if len(subject) >= 36:
            separator = "**"

        hours += hour + separator
        marks += mark + separator
    

    diploms.append({
        "Фамилия": first,
        "Имя": middle,
        "Отчество": last,
        "число_месяцполностью_год__рождения": birthdate,
        "Председатель": predsedatel,
        "Дата_выдачи": data_vidachi,
        "Дисциплины": subjects,
        "Часы": hours,
        "оценки": marks,
        "Предыдущий_документ": pred_document,
        "Профессия": svid_prof,
        "Проф_квалиф": svid_qual,
        "Проф_номер": svid_nom_obl,
        "Проф_рег_н": svid_reg_nom,
        "Проф_дисц": subjects,
        "Проф_часы": hours,
        "Проф_оцен": marks,
        "Проф_решен": svid_resh
    })

with MailMerge('шаблон_свид_обло.docx', options=mailmerge_options) as document:
    document.merge_templates(diploms, separator='continuous_section')
    document.write('свид_обло.docx')

with MailMerge('шаблон_свид_прил.docx', options=mailmerge_options) as document:
    document.merge_templates(diploms, separator='continuous_section')
    document.write('свид_прил.docx')