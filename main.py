from openpyxl import load_workbook
from mailmerge import MailMerge, MailMergeOptions, OptionAutoUpdateFields, OptionKeepFields
import qrcode
from io import BytesIO
import tempfile
import os
from PIL import Image
from pathlib import Path
import warnings
from docx import Document
from docx.oxml.ns import qn
from lxml import etree
import copy
from datetime import datetime
import locale
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def convertMarktoWord(mark):
    match mark:
        case "5":
            return "отлично"
        case "4":
            return "хорошо"
        case "3":
            return "удовлетворительно"
        case "x":
            return "x"
        case _:
            return mark


def fix_encoding(text):
    """Исправляет кодировку Windows-1251"""
    if not text:
        return ""
    
    # Если строка содержит последовательности вида РўРµСЃР»РµРЅРєРѕ
    if re.search(r'Р[А-Я]', str(text)):
        try:
            # Пробуем декодировать из Windows-1251
            return str(text).encode('latin-1').decode('windows-1251')
        except:
            return str(text)
    return str(text)


def format_date_ru(date_str):
    # Парсим исходную дату

    # Словарь с русскими названиями месяцев в родительном падеже
    months = {
        1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
        5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
        9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
    }

    # Форматируем: день с ведущим нулём, месяц из словаря, год, слово "года"
    return f"{date_str.day:02d} {months[date_str.month]} {date_str.year} года"


def get_srok(attestat, common_data):
    if attestat == "Аттестат об основном общем образовании":
        return common_data["srok9"]
    else:
        return common_data["srok11"]


def getQrPath(i, fio, data_vidachi, seria, nomer, demo_kod, demo_type, max_point, demo_point):
    demo_data = []
    common = [fio, data_vidachi, seria  + " " + nomer]
    if (demo_point != ""):
        demo_result = "результат: " + demo_point + " из " + str(max_point) + " баллов"
        demo_data.append(demo_type)
        demo_data.append(demo_kod)
        demo_data.append(demo_result)
    
    qr_path = " | ".join(common + demo_data)
    qr = qrcode.make(fix_encoding(qr_path))
    filename = f'img/qr_{i}.png'
    qr.save(filename)
    print(fix_encoding(qr_path))
    return Path(filename).absolute().as_posix()


mailmerge_options = MailMergeOptions(
    remove_empty_tables=False,
    auto_update_fields_on_open=OptionAutoUpdateFields.ALWAYS,
    # keep_fields=OptionKeepFields.NONE,
    # merge_if_fields=True,
    table_rows_replace_mode=False)


data_vidachi = "01 июля 2026 года"
gruppa = "ПСД-35"
diploms = []
wb = load_workbook(r'C:\\Users\\ganz\\OneDrive\\Документы\\колледж onedrive\\ведомости\\' + gruppa + '.xlsx')
wb_common = load_workbook('utils.xlsx', read_only=True)
ws_common = wb_common["data"]
ws_data = wb["Сведения"]
ws_mark = wb["Оценки"]
common = {}

for row in ws_common.values:
    grupp_name = row[0]
    if grupp_name == gruppa:
        common = {
            "speciality": row[1],
            "reshenie": row[2],
            "predsedatel": row[3],
            "srok9": row[4],
            "srok11": row[5],
            "vidacha": row[6],
            "qualification" : row[7],
            "demo_kod": row[8],
            "demo_type": row[9],
            "demo_max_point": row[10],
            "up-vd": row[11],
            "up-so": row[12],
            "up-mp": row[13],
            "pp-vd": row[14],
            "pp-so": row[15],
            "svid-qual": row[16],
            "svid-prof": row[17],
        }
        # Вывод колонок A и C
for i, row in enumerate(ws_data.iter_rows(min_row=2, max_row=50, min_col=1, max_col=13, values_only=True)):
    fio = row[0]
    if (fio == None): # Если дойти до пустой строки, то остановить цикл
        break

    first, middle, last = fio.split(' ')
    dt_obj = row[2]
    birthdate = format_date_ru(row[2])
    red = row[5]
    reg = row[1]
    attestat_type = str(row[3])
    attestat_year = str(row[4])
    pred_document = attestat_type + ", " + attestat_year + " года"
    srok = get_srok(attestat_type, common)
    qualification = common["qualification"]
    speciality = common["speciality"]
    predsedatel = common["predsedatel"]
    reshenie = format_date_ru(common["reshenie"])
    tema_vkr = row[6]
    seria = str(row[7])
    nomer = str(row[8])
    up_vd = common["up-vd"]
    up_so = common["up-so"]
    up_mp = common["up-mp"]
    pp_vd = common["up-vd"]
    pp_so = common["up-so"]
    pp_mp = row[10]
    svid_resh = reshenie
    svid_qual = common["svid-qual"]
    svid_prof = common["svid-prof"]
    svid_reg_nom = row[11]
    svid_nom_obl = row[12]
    svid_subjects = ""
    svid_hours = ""
    svid_marks = ""
    
    subjects = ""
    hours = ""
    marks = ""

    kursovie_subjects = ""
    kursovie_marks = ""

    demo_point = ""

    student_col = i+3
    for mark_col in ws_mark.iter_rows(min_row=2, max_row=100, min_col=0, max_col=student_col, values_only=True):
        subject = mark_col[0]
        hour = mark_col[1]
        mark = mark_col[student_col-1]

        if subject == None:
            break
        if mark == None:
            continue

        subject = str(subject)
        hour = str(hour)
        mark = str(mark).replace(".",",")

        mark = convertMarktoWord(mark) # Преобразовать числовые оценки в "отлично", "хорошо" и тд.
        
        if subject == common["demo_kod"]:
            demo_point = mark

        separator = "*"
        
        if subject == "Производственная практика" or "Делопроизводитель" in subject:
            svid_subjects += subject + separator * 2
            svid_marks += mark + separator * 4

            if subject == "Производственная практика":
                svid_hours += "2 недели"
            else:
                svid_hours += hour + separator * 4

        if "ВСЕГО часов" in subject or "в том числе аудиторных:" in subject:
            if "|" in hour:
                hour9, hour11 = hour.split("|")

                if attestat_type == "Аттестат о среднем общем образовании":
                    hour = hour11
                else:
                    hour = hour9
        


        if ("Курсовая" in subject):
            kursovie_subjects += subject + separator
            kursovie_marks += mark + separator
        else:
            subjects += subject + separator

            if len(subject) >= 90:
                separator = "**"

            hours += hour + separator
            marks += mark + separator
    
    if(tema_vkr != None):
        subjects += tema_vkr

    file = getQrPath(i, fio, common["vidacha"], seria, nomer, common["demo_kod"], common["demo_type"], common["demo_max_point"], demo_point)

    diploms.append({
        "Фамилия": first,
        "Имя": middle,
        "Отчество": last,
        "Регистр_номер": reg,
        "число_месяцполностью_год__рождения": birthdate,
        "Решение_госуд_Экз_комисии": reshenie,
        "Председатель": predsedatel,
        "Дата_выдачи": data_vidachi,
        "Специальность": speciality,
        "Красный": red,
        "file": file,
        "Дисциплины": subjects,
        "Часы": hours,
        "оценки": marks,
        "Предыдущий_документ": pred_document,
        "срок": srok,
        "Квалификация": qualification,
        "Специальность": speciality,
        "курсовые": kursovie_subjects,
        "оценка_курсовая": kursovie_marks,
        "уп_вид_д": up_vd,
        "уп_исп_ср": up_so,
        "уп_место": up_mp,
        "пп_вид_д": pp_vd,
        "пп_исп_ср": pp_so,
        "пп_место": pp_mp,
        "Профессия": svid_prof,
        "Проф_квалиф": svid_qual,
        "Проф_номер": svid_nom_obl,
        "Проф_рег_н": svid_reg_nom,
        "Проф_дисц": svid_subjects,
        "Проф_часы": svid_hours,
        "Проф_оцен": svid_marks,
        "Проф_решен": svid_resh
    })

with MailMerge('cover_oblo.docx', options=mailmerge_options) as document:
    document.merge_templates(diploms, separator='continuous_section')
    document.write('diploms_oblo.docx')

with MailMerge('cover_prilo.docx', options=mailmerge_options) as document:
    document.merge_templates(diploms, separator='continuous_section')
    document.write('diploms_prilo.docx')

with MailMerge('шаблон_свид_обло.docx', options=mailmerge_options) as document:
    document.merge_templates(diploms, separator='continuous_section')
    document.write('свид_обло.docx')

with MailMerge('шаблон_свид_прил.docx', options=mailmerge_options) as document:
    document.merge_templates(diploms, separator='continuous_section')
    document.write('свид_прил.docx')